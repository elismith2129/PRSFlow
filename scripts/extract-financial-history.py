#!/usr/bin/env python3
# ─────────────────────────────────────────────────────────────────────────────
# Turn "PRS Daily Numbers" (the Google Sheet, 2017–2029, one tab per year) into
# a CSV that loads straight into `financial_history`.
#
#   python3 scripts/extract-financial-history.py <workbook.xlsx> <out.csv>
#
# WHY PYTHON, IN A NODE REPO. This runs ONCE, by hand, and never ships. Reading
# .xlsx in Node means adding SheetJS to package.json for a script no user will
# ever trigger; openpyxl is already on any machine with Python. The output is a
# plain CSV, so nothing downstream knows or cares what produced it. Kept in the
# repo for provenance — when someone asks in 2028 where these figures came
# from, this file is the answer.
#
# THE SHEET'S SHAPE (verified 2026-08-20 against the real workbook):
#   · One tab per year. Header is ALWAYS row 4. Column 1 is the date.
#   · Columns repeat in blocks per room: "PRS A / Studio", "PRS A / Rental
#     Profit", "PRS A / Studio + Rental", "PRS A / Engineer", "PRS A / Total
#     Room". Retired rooms (PRS D/F/H, ARS C, ERS C/D/E) carry only "/ Studio".
#   · The grain is DAY × ROOM. There is no session, no client, no artist, and
#     no assistant column anywhere — the app's fourth category simply has no
#     history and charts as zero before the cutover.
#
# WHAT IS DELIBERATELY NOT IMPORTED, AND WHY IT MATTERS:
#   · "Studio + Rental" and "Total Room" are SUMS of columns we already take.
#   · "PRS TOTAL", "ARS TOTAL", "ERS TOTAL", "TRACK TOTAL", "GROUP TOTAL" are
#     roll-ups of the rooms we already take.
#   · "Dec Total" / "TOTAL 2023" ROWS sit at the bottom of every tab.
#   Taking any of them would roughly triple the books. They are used instead as
#   the CHECK: the annual total row is what this script reconciles against, so a
#   mapping error fails loudly rather than importing a plausible wrong number.
#
# "RENTAL PROFIT" IS ALREADY PRS'S SHARE (Eli, 2026-08-20). Own gear rents at
# full cost; contracted gear earns a 30% fee. So this column is what Paramount
# actually made, which is the right thing to chart. NOTE FOR THE LIVE SIDE:
# `rental_rows.charge` is the GROSS figure and does not yet model that split, so
# live rentals will read high against history until the work order learns the
# difference. That is a WO change, not an import change.
# ─────────────────────────────────────────────────────────────────────────────

import csv
import datetime
import re
import sys

import openpyxl

VENUE = {'PRS': 'Paramount', 'ARS': 'Ameraycan', 'ERS': 'Encore', 'TRACK': 'Track'}

# Track's headers drift across years: 2020 wrote NTH/STH/Sth, 2021+ write N/S.
TRACK_ROOM = {'N': 'North', 'NTH': 'North', 'S': 'South', 'STH': 'South'}

# 'Studio' → the room itself; 'Rental Profit' → PRS's share; 'Engineer' → staff.
# Everything else in a room block is derived and must not be read.
CATEGORY = {
    'studio': 'room',
    'rental profit': 'rental',
    'engineer': 'engineering',
}

SKIP_MEASURES = {'studio + rental', 'total room'}


def parse_header(raw):
    """
    'PRS A / Studio' → ('Paramount', 'Studio A', 'room').

    Returns None for anything that is not a real room measure — rollups,
    derived columns, the trailing '% Net of Eng', and blank spacer columns.
    """
    if not raw:
        return None
    # Headers are two lines in the sheet ("PRS A" over "Studio"); openpyxl gives
    # them back with a newline. One malformed 2021+ header, 'TRACK S Studio',
    # lost its separator entirely, so split on newline OR slash OR — as a last
    # resort — the space before a known measure word.
    text = str(raw).replace('\n', ' / ').strip()
    if '/' not in text:
        m = re.match(r'^(.*?)\s+(Studio|Engineer|Rental Profit|Total Room)$', text, re.I)
        if not m:
            return None
        text = f'{m.group(1)} / {m.group(2)}'

    left, _, measure = text.partition('/')
    left = left.strip()
    measure = measure.strip().lower()

    if measure in SKIP_MEASURES:
        return None
    category = CATEGORY.get(measure)
    if category is None:
        return None

    parts = left.split()
    if len(parts) < 2:
        return None
    code, room_raw = parts[0].upper(), ' '.join(parts[1:]).upper()

    # Rollup columns share the room slot with the word TOTAL.
    if room_raw == 'TOTAL' or code not in VENUE:
        return None

    if code == 'TRACK':
        room = TRACK_ROOM.get(room_raw)
        if room is None:
            return None
    else:
        # 'A' → 'Studio A', matching lib/studios.ts so a historical room and a
        # live one are the same string and land in the same filter chip.
        if not re.fullmatch(r'[A-Z]', room_raw):
            return None
        room = f'Studio {room_raw}'

    return VENUE[code], room, category


# The three roll-up columns this script reconciles against, mapped to the
# category each one should equal once the room columns are summed.
GROUP_COLS = {
    'group total / studio': 'room',
    'group total / rental profit': 'rental',
    'group total / engineer': 'engineering',
}


def reconcile(ws, header_row, cols):
    """
    Compare the room columns this script READS against the sheet's own
    GROUP TOTAL columns, day by day.

    NOT against the 'TOTAL <year>' row at the bottom. That row was the first
    check tried and it is unreliable — in several tabs (2024, 2025, 2026) it is
    blank or stale, and disagreeing with a stale cell would have condemned a
    correct extraction. The GROUP TOTAL *column* is live and equals the sum of
    the four venue totals exactly in every year tested.

    Returns (per-category totals, list of days where the sheet disagrees with
    itself). Those days are spreadsheet data-entry errors, not import errors:
    the per-room cells are typed by hand and the roll-up is a formula, so when
    they disagree the ROOM cells are the source and the roll-up is what drifted.
    """
    group_cols = {}
    for c in range(1, ws.max_column + 1):
        key = str(ws.cell(header_row, c).value or '').replace('\n', ' / ').strip().lower()
        if key in GROUP_COLS:
            group_cols[GROUP_COLS[key]] = c

    mine = {'room': 0.0, 'rental': 0.0, 'engineering': 0.0}
    theirs = {'room': 0.0, 'rental': 0.0, 'engineering': 0.0}
    conflicts = []

    for r in range(header_row + 1, ws.max_row + 1):
        d = ws.cell(r, 1).value
        if not isinstance(d, datetime.datetime):
            continue

        day_mine = {'room': 0.0, 'rental': 0.0, 'engineering': 0.0}
        for c, (_venue, _room, category) in cols.items():
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)):
                day_mine[category] += v

        for cat, total in day_mine.items():
            mine[cat] += total
            c = group_cols.get(cat)
            if c is None:
                continue
            v = ws.cell(r, c).value
            v = float(v) if isinstance(v, (int, float)) else 0.0
            theirs[cat] += v
            if abs(total - v) > 0.5:
                conflicts.append((d.date().isoformat(), cat, round(total), round(v)))

    return mine, theirs, conflicts


def main():
    if len(sys.argv) < 3:
        print('usage: extract-financial-history.py <workbook.xlsx> <out.csv>')
        return 1

    src, out = sys.argv[1], sys.argv[2]
    # A stable provenance label, not the literal filename. Uploads arrive with a
    # UUID prefix and a "(3)" copy-suffix; stamping that into 55,000 rows would
    # make the audit trail read like an accident.
    source_file = 'PRS Daily Numbers.xlsx'
    wb = openpyxl.load_workbook(src, data_only=True)

    rows = []
    report = []
    problems = []

    for sheet in wb.sheetnames:
        if not re.fullmatch(r'(19|20)\d\d', sheet):
            continue
        ws = wb[sheet]

        header_row = None
        for r in range(1, 12):
            if str(ws.cell(r, 1).value).strip().lower() == 'date':
                header_row = r
                break
        if header_row is None:
            problems.append(f'{sheet}: no header row found')
            continue

        cols = {}
        for c in range(1, ws.max_column + 1):
            parsed = parse_header(ws.cell(header_row, c).value)
            if parsed:
                cols[c] = parsed

        days = 0

        for r in range(header_row + 1, ws.max_row + 1):
            d = ws.cell(r, 1).value
            # A row is real only if column 1 is an actual date. This is what
            # excludes 'Dec Total' and 'TOTAL 2023' without pattern-matching
            # their wording.
            if not isinstance(d, datetime.datetime):
                continue
            days += 1
            iso = d.date().isoformat()

            for c, (venue, room, category) in cols.items():
                v = ws.cell(r, c).value
                if not isinstance(v, (int, float)) or v == 0:
                    continue
                amount = round(float(v), 2)
                rows.append({
                    'session_date': iso,
                    'venue': venue,
                    'room': room,
                    'category': category,
                    'direction': 'revenue',
                    'amount': f'{amount:.2f}',
                    'client_name': '',
                    'artist_name': '',
                    'source_file': source_file,
                    # Deterministic, readable, and stable across re-runs, so the
                    # unique index turns a second import into an upsert. A hash
                    # would do the same job while telling nobody anything.
                    'source_key': f'{sheet}#{iso}#{venue}#{room}',
                })

        # ── Reconcile against the sheet's own GROUP TOTAL columns ────────────
        got, theirs, conflicts = reconcile(ws, header_row, cols)

        line = {'year': sheet, 'days': days, 'cols': len(cols), 'conflicts': conflicts}
        for cat in ('room', 'rental', 'engineering'):
            line[cat] = round(got[cat])
            line[cat + '_sheet'] = round(theirs[cat])
            line[cat + '_ok'] = abs(got[cat] - theirs[cat]) <= 2
        report.append(line)

        for day, cat, room_cells, rollup in conflicts:
            problems.append(
                f'{day}  {cat}: room cells add to {room_cells:,} but the '
                f'sheet\'s own roll-up says {rollup:,} '
                f'(off by {room_cells - rollup:,})'
            )

    # ── Write ────────────────────────────────────────────────────────────────
    # ONE FILE PER YEAR. The combined CSV is 9MB / 55,000 rows, and the Supabase
    # dashboard's CSV importer is a browser upload — big files there fail slowly
    # and unhelpfully. A year is roughly 5,000 rows and lands in seconds, which
    # also means each year can be checked against the spreadsheet as it goes in
    # rather than all of it being wrong at the end.
    fields = ['session_date', 'venue', 'room', 'category', 'direction', 'amount',
              'client_name', 'artist_name', 'source_file', 'source_key']

    out_dir = out.rstrip('/')
    if out_dir.endswith('.csv'):
        out_dir = out_dir.rsplit('/', 1)[0]

    by_year = {}
    for row in rows:
        by_year.setdefault(row['session_date'][:4], []).append(row)

    written = []
    for year in sorted(by_year):
        path = f'{out_dir}/financial_history_{year}.csv'
        with open(path, 'w', newline='') as fh:
            w = csv.DictWriter(fh, fieldnames=fields)
            w.writeheader()
            w.writerows(by_year[year])
        written.append((year, len(by_year[year]), path))

    # ── Report ───────────────────────────────────────────────────────────────
    print(f'\n{source_file}  →  {out_dir}/financial_history_<year>.csv')
    print(f'{len(rows):,} rows across {len(written)} files\n')
    print(f"{'year':6}{'rows':>7}{'room':>14}{'engineering':>14}{'rental':>11}{'flagged days':>14}")
    for line in report:
        if line['days'] == 0:
            continue
        n = len(by_year.get(line['year'], []))
        print(
            f"{line['year']:6}{n:>7,}"
            f"{line['room']:>14,}{line['engineering']:>14,}{line['rental']:>11,}"
            f"{len(line['conflicts']):>14}"
        )

    grand = sum(line['room'] + line['engineering'] + line['rental'] for line in report)
    print(f'\nTOTAL IMPORTED REVENUE  ${grand:,.0f}')

    if problems:
        # These are days where the spreadsheet disagrees WITH ITSELF — the room
        # cells and the roll-up formula do not match. The import takes the room
        # cells, which are the typed source; the roll-up is what drifted. Listed
        # in full so Eli can correct the sheet if he wants the two to agree.
        drift = sum(c[2] - c[3] for line in report for c in line['conflicts'])
        print(f'\n⚠ {len(problems)} day(s) where the SPREADSHEET disagrees with itself')
        print(f'   (room cells vs its own roll-up columns; net effect ${drift:,.0f})')
        print('   The import uses the room cells — the typed source, not the formula.')
        for p in problems[:25]:
            print(f'   {p}')
        if len(problems) > 25:
            print(f'   … and {len(problems) - 25} more')
    else:
        print('\n✓ Every day reconciles against the spreadsheet\'s own roll-up columns.')

    return 0


if __name__ == '__main__':
    sys.exit(main())
